import openpyxl
import requests
import os

# Google Sheets details
spreadsheet_id = "12m4Sq4CMnLB9nn6INxKUdSNSIlTJes_uTbg19RE9-X0"

# Construct the export URL for .xlsx format
xlsx_export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"

# Output file
output_file = "mission_database.xlsx"

try:
    # Send GET request
    response = requests.get(xlsx_export_url)

    # Check if the request was successful
    if response.status_code == 200:
        # Write the content to a file
        with open(output_file, "wb") as file:
            file.write(response.content)
        print(f"Excel file downloaded successfully as '{output_file}'")
    else:
        print(f"Failed to download Excel file. HTTP Status Code: {response.status_code}")
except Exception as e:
    print(f"An error occurred: {e}")

# Type encoding dictionary for the last 4 bits
type_encoding = {
    "uint8_t": 0,
    "bool_t" : 0,
    "int8_t": 1,
    "uint16_t": 2,
    "int16_t": 3,
    "uint32_t": 4,
    "int32_t": 5,
    "uint64_t": 6,
    "int64_t": 7,
    "float_t": 8,
    "double_t": 9,
}

# Function to encode the ID
def encode_id(numeric_id, variable_type):
    if variable_type not in type_encoding:
        print(f"Error: Type '{variable_type}' not found in type encoding dictionary. Defaulting to 'uint64_t'.")
        variable_type = "uint64_t"
    type_code = type_encoding[variable_type]
    return (numeric_id << 4) | type_code

# Subsystem acronyms and their corresponding number offsets
subsystem_config = {
    "OBDH": 5000,
    "COMMS": 10000,
    "PAY": 15000,
    "ADCS": 20000,
    "EPS": 25000,
}

# Dictionary mapping C++ types to _t counterparts
cpp_type_map = {
    "bool": "bool_t",
    "float": "float_t",
    "double": "double_t",
    "int": "int_t",
    "uint8_t": "uint8_t",
    "uint16_t": "uint16_t",
    "uint32_t": "uint32_t",
    "uint64_t": "uint64_t",
    "int8_t": "int8_t",
    "int16_t": "int16_t",
    "int32_t": "int32_t",
    "int64_t": "int64_t",
    "char": "char_t",
    "enum": "enum",
}

# Paths to the Excel file and output directories
excel_file = "mission_database.xlsx"
output_dir = "src/main/yamcs/mdb/peaksat/"

# Ensure the output directories exist
os.makedirs(output_dir, exist_ok=True)

# Set to track unique IDs
processed_ids = set()

# Load the workbook
workbook = openpyxl.load_workbook(excel_file)

# Lists to collect output lines
dt_lines = []
xtce_lines = []

# Add the header lines for the xtce file
xtce_lines.append('<?xml version="1.0" encoding="UTF-8"?>')
xtce_lines.append('<SpaceSystem name="peaksat-xtce" xmlns:xtce="http://www.omg.org/spec/XTCE/20180204"')
xtce_lines.append('    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"')
xtce_lines.append('    xsi:schemaLocation="http://www.omg.org/spec/XTCE/20180204 https://www.omg.org/spec/XTCE/20180204/SpaceSystem.xsd"')
xtce_lines.append('    shortDescription="This is a bogus satellite telemetry and telecommand database."')
xtce_lines.append('    operationalStatus="unittest">')
xtce_lines.append(' ')
xtce_lines.append('    <xtce:TelemetryMetaData>')
xtce_lines.append('        <ParameterSet>')
xtce_lines.append('            <!--OBC/OBDH Parameters-->')

# TODO: Add the header lines for the dt file
dt_lines.append('<?xml version="1.0" encoding="UTF-8"?>')
dt_lines.append('<SpaceSystem name="peaksat-dt" xmlns:xtce="http://www.omg.org/spec/XTCE/20180204"')
dt_lines.append('    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"')
dt_lines.append('    xsi:schemaLocation="http://www.omg.org/spec/XTCE/20180204 https://www.omg.org/spec/XTCE/20180204/SpaceSystem.xsd"')
dt_lines.append('    shortDescription="This is a bogus satellite telemetry and telecommand database."')
dt_lines.append('    operationalStatus="unittest">')
dt_lines.append(' ')
dt_lines.append('    <!-- Contains PeakSats\'s non primitive parameters and arguments -->')
dt_lines.append(' ')
dt_lines.append('    <xtce:TelemetryMetaData>')
dt_lines.append('        <ParameterTypeSet>')

parameter_id_lines = []
parameter_id_lines.append('            <EnumeratedArgumentType name="parameterId_t">')
parameter_id_lines.append('                <IntegerDataEncoding sizeInBits="16" />')
parameter_id_lines.append('                <EnumerationList>')

# Process each subsystem separately
namespace_blocks = {acronym: [] for acronym in subsystem_config.keys()}
valid_rows = []

# Collect all valid rows across all sheets
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        id_cell = row[0]  # First column
        variable_cell = row[4]  # Fifth column

        # Skip rows with no variable name
        if not variable_cell.value or not isinstance(variable_cell.value, str):
            continue

        # Check if the first column contains a valid ID
        if id_cell.value and isinstance(id_cell.value, str):
            # Add to valid rows
            valid_rows.append(row)

# Process all valid rows
for idx, row in enumerate(valid_rows):
    id_cell = row[0]  # First column
    type_cell = row[2]  # Third column
    variable_cell = row[4]  # Fifth column
    enum_items_cell = row[6]  # Seventh column
    value_cell = row[7]  # Eighth column

    # Identify the subsystem acronym
    for acronym, offset in subsystem_config.items():
        if id_cell.value.startswith(f"{acronym}-"):
            # Extract the numeric part of the ID
            numeric_id = id_cell.value[len(acronym) + 1:]  # Remove acronym and '-'

            try:
                numeric_id = int(numeric_id) + offset
            except ValueError:
                print(f"Skipping invalid numeric ID: {id_cell.value}")
                continue

            # Skip duplicates
            if numeric_id in processed_ids:
                continue
            processed_ids.add(numeric_id)

            # Get variable name and type
            variable_name = variable_cell.value.strip()

            ### TODO: Fix special case handling
            variable_type = cpp_type_map[type_cell.value.strip()] if type_cell.value and type_cell.value.strip() in cpp_type_map else "uint32_t"
            
            enum_items = enum_items_cell.value.strip() if enum_items_cell.value else ""
            # Handle float values to remove .0 for whole numbers
            if value_cell.value:
                if isinstance(value_cell.value, float) and value_cell.value.is_integer():
                    param_value = str(int(value_cell.value))  # Convert to int if it's a whole number
                else:
                    param_value = str(value_cell.value).strip()  # Convert to string for other cases
            else:
                param_value = "0"

            if acronym == "EPS":
                acronym = ""

            parameter_lines = []
            if variable_type == "enum":
                parameter_lines.append(f"            <Parameter parameterTypeRef=\"peaksat-dt/{variable_name}_t\" name=\"{acronym}{variable_name}\"")
            else:
                parameter_lines.append(f"            <Parameter parameterTypeRef=\"base-dt/{variable_type}\" name=\"{acronym}{variable_name}\"")

            # Add to the corresponding namespace block
            # block_lines = namespace_blocks[acronym]
            # block_lines.append(f"        {variable_name}ID = {numeric_id}")

            #### Handle special cases
            if variable_type == f"{variable_name}_t":
                variable_type = "enum"

            # Enum definitions (if type is "enum")
            enum_lines = []
            if variable_type == "enum":
                enum_lines.append(f'')
                enum_lines.append(f'            <EnumeratedParameterType name="{variable_name}_t">')
                enum_lines.append('                <IntegerDataEncoding sizeInBits="8" />')
                enum_lines.append('                <EnumerationList>')
                for enum_item in enum_items.split(","):
                    enum_item = enum_item.strip()
                    if not enum_item:
                        continue
                    enum_item_parts = enum_item.split(" = ")
                    if len(enum_item_parts) == 2:
                        label, value = enum_item_parts
                        enum_lines.append(f'                    <Enumeration label="{label.strip()}" value="{value.strip()}" />')
                enum_lines.append('                </EnumerationList>')
                enum_lines.append('            </EnumeratedParameterType>')
                enum_lines.append(f'')

            dt_lines.append("\n".join(enum_lines))

            # Encode the ID with the new encoding rule
            encoded_id = encode_id(numeric_id, variable_type)

            # Skip duplicates
            if encoded_id in processed_ids:
                continue


            xtce_lines.append(" ".join(parameter_lines) + "/>")

            # Create the parameter ID line
            parameter_id_lines.append(f'                    <Enumeration value=\"{encoded_id}\" label=\"{acronym}{variable_name}\" />')

            break

# Finalize the xtce file
#     dt_lines.append('        </ParameterTypeSet>')
            #     dt_lines.append('    </xtce:TelemetryMetaData>')
xtce_lines.append('        </ParameterSet>\n')
xtce_lines.append('    </xtce:TelemetryMetaData>\n')
xtce_lines.append('</SpaceSystem>')

# Finalize the dt file
dt_lines.append('        </ParameterTypeSet>')
dt_lines.append('    </xtce:TelemetryMetaData>')
dt_lines.append('</SpaceSystem>')

id_lines = []

id_lines.append('<!--Contains all ParameterTypes for Telemetry and ArgumentTypes for Telecommanding.-->')
id_lines.append('    <TelemetryMetaData>')
id_lines.append('        <ParameterTypeSet>')
id_lines.append('            <EnumeratedParameterType name="parameterId_t">')
id_lines.append('                <IntegerDataEncoding sizeInBits="16" />')
id_lines.append('                <EnumerationList>')
id_lines.append(parameter_id_lines)
id_lines.append('                </EnumerationList>')
id_lines.append('            </EnumeratedParameterType>')
id_lines.append('')
id_lines.append('            <ArrayParameterType name="parameterIdArray_t" arrayTypeRef="parameterId_t">')
id_lines.append('               <DimensionList>')
id_lines.append('                   <Dimension>')
id_lines.append('                       <StartingIndex>')
id_lines.append('                           <FixedValue>0</FixedValue>')
id_lines.append('                      </StartingIndex>')
id_lines.append('                      <EndingIndex>')
id_lines.append('                           <DynamicValue>')
id_lines.append('                               <ArgumentInstanceRef argumentRef="total_parameters" />')
id_lines.append('                               <LinearAdjustment intercept="-1" />')
id_lines.append('                           </DynamicValue>')
id_lines.append('                       </EndingIndex>')
id_lines.append('                   </Dimension>')
id_lines.append('               </DimensionList>')
id_lines.append('           </ArrayParameterType>')
id_lines.append('        </ParameterTypeSet>')
id_lines.append('    </TelemetryMetaData>')
id_lines.append('')
id_lines.append('    <CommandMetaData>')
id_lines.append('        <ArgumentTypeSet>')
id_lines.append('            <EnumeratedArgumentType name="parameterId_t">')
id_lines.append('                <IntegerDataEncoding sizeInBits="16" />')
id_lines.append('                <EnumerationList>')
id_lines.append(parameter_id_lines)
id_lines.append('                </EnumerationList>')
id_lines.append('            </EnumeratedArgumentType>')
id_lines.append('        </ArgumentTypeSet>')
id_lines.append('')
id_lines.append('       <ArgumentTypeSet>')
id_lines.append('           <ArrayArgumentType name="parameterIdArray_t" arrayTypeRef="parameterId_t">')
id_lines.append('              <DimensionList>')
id_lines.append('                  <Dimension>')
id_lines.append('                      <StartingIndex>')
id_lines.append('                          <FixedValue>0</FixedValue>')
id_lines.append('                     </StartingIndex>')
id_lines.append('                     <EndingIndex>')
id_lines.append('                          <DynamicValue>')
id_lines.append('                              <ArgumentInstanceRef argumentRef="total_parameters" />')
id_lines.append('                              <LinearAdjustment intercept="-1" />')
id_lines.append('                          </DynamicValue>')
id_lines.append('                      </EndingIndex>')
id_lines.append('                  </Dimension>')
id_lines.append('              </DimensionList>')
id_lines.append('          </ArrayArgumentType>')
id_lines.append('      <ArgumentTypeSet>')
id_lines.append('   <CommandMetaData>')
id_lines.append('</SpaceSystem>')

# Write the xtce file
output_xtce_file = f"{output_dir}peaksat-xtce.xml"
output_dt_file = f"{output_dir}peaksat-dt.xml"
output_parameter_id_file = f"src/main/yamcs/mdb/peaksat/parameter_id.xml"

print(parameter_id_lines)

with open(output_xtce_file, "w") as xtce_file:
    xtce_file.write("\n".join(xtce_lines))

with open(output_dt_file, "w") as dt_file:
    dt_file.write("\n".join(line for line in dt_lines if line.strip()))

# Flatten the parameter_id_lines list
id_lines = [item for sublist in id_lines for item in (sublist if isinstance(sublist, list) else [sublist])]

with open(output_parameter_id_file, "w") as parameter_id_file:
    parameter_id_file.write("\n".join(id_lines))

print(f"Processing complete.")
print(f"Generated XTCE file: {output_dir}obc-xtce.xml")
print(f"Generated DT file: {output_dir}obc-dt.xml")