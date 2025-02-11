import openpyxl
import requests
import os

def download_excel_file(url, output_file):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            with open(output_file, "wb") as file:
                file.write(response.content)
            print(f"Excel file downloaded successfully as '{output_file}'")
        else:
            print(f"Failed to download Excel file. HTTP Status Code: {response.status_code}")
    except Exception as e:
        print(f"An error occurred: {e}")

def encode_id(numeric_id, variable_type, type_encoding):
    if variable_type not in type_encoding:
        print(f"Error: Type '{variable_type}' not found in type encoding dictionary. Defaulting to 'uint64_t'.")
        variable_type = "uint64_t"
    type_code = type_encoding[variable_type]
    return (numeric_id << 4) | type_code

def process_excel_file(excel_file, subsystem_config, cpp_type_map, type_encoding):
    workbook = openpyxl.load_workbook(excel_file)
    processed_ids = set()
    valid_rows = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            id_cell = row[0]
            variable_cell = row[4]
            if not variable_cell.value or not isinstance(variable_cell.value, str):
                continue
            if id_cell.value and isinstance(id_cell.value, str):
                valid_rows.append(row)

    dt_lines, xtce_lines, parameter_id_lines = initialize_output_lines()

    for idx, row in enumerate(valid_rows):
        id_cell = row[0]
        type_cell = row[2]
        variable_cell = row[4]
        enum_items_cell = row[6]
        value_cell = row[7]

        for acronym, offset in subsystem_config.items():
            if id_cell.value.startswith(f"{acronym}-"):
                numeric_id = id_cell.value[len(acronym) + 1:]
                try:
                    numeric_id = int(numeric_id) + offset
                except ValueError:
                    print(f"Skipping invalid numeric ID: {id_cell.value}")
                    continue

                if numeric_id in processed_ids:
                    continue
                processed_ids.add(numeric_id)

                variable_name = variable_cell.value.strip()
                variable_type = cpp_type_map[type_cell.value.strip()] if type_cell.value and type_cell.value.strip() in cpp_type_map else "uint32_t"
                enum_items = enum_items_cell.value.strip() if enum_items_cell.value else ""
                param_value = str(int(value_cell.value)) if value_cell.value and isinstance(value_cell.value, float) and value_cell.value.is_integer() else str(value_cell.value).strip() if value_cell.value else "0"

                if acronym == "EPS":
                    acronym = ""

                parameter_lines, enum_lines = generate_parameter_lines(variable_name, variable_type, enum_items, acronym)
                dt_lines.append("\n".join(enum_lines))

                encoded_id = encode_id(numeric_id, variable_type, type_encoding)
                if encoded_id in processed_ids:
                    continue

                xtce_lines.append(" ".join(parameter_lines) + "/>")
                parameter_id_lines.append(f'                    <Enumeration value=\"{encoded_id}\" label=\"{acronym}{variable_name}\" />')
                break

    finalize_output_files(dt_lines, xtce_lines, parameter_id_lines)

def initialize_output_lines():
    dt_lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<SpaceSystem name="peaksat-dt" xmlns:xtce="http://www.omg.org/spec/XTCE/20180204"',
        '    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"',
        '    xsi:schemaLocation="http://www.omg.org/spec/XTCE/20180204 https://www.omg.org/spec/XTCE/20180204/SpaceSystem.xsd"',
        '    shortDescription="This is a bogus satellite telemetry and telecommand database."',
        '    operationalStatus="unittest">',
        ' ',
        '    <!-- Contains PeakSats\'s non primitive parameters and arguments -->',
        ' ',
        '    <xtce:TelemetryMetaData>',
        '        <ParameterTypeSet>'
    ]

    xtce_lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<SpaceSystem name="peaksat-xtce" xmlns:xtce="http://www.omg.org/spec/XTCE/20180204"',
        '    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"',
        '    xsi:schemaLocation="http://www.omg.org/spec/XTCE/20180204 https://www.omg.org/spec/XTCE/20180204/SpaceSystem.xsd"',
        '    shortDescription="This is a bogus satellite telemetry and telecommand database."',
        '    operationalStatus="unittest">',
        ' ',
        '    <xtce:TelemetryMetaData>',
        '        <ParameterSet>',
        '            <!--OBC/OBDH Parameters-->'
    ]

    parameter_id_lines = []

    return dt_lines, xtce_lines, parameter_id_lines

def generate_parameter_lines(variable_name, variable_type, enum_items, acronym):
    parameter_lines = []
    if variable_type == "enum":
        parameter_lines.append(f"            <Parameter parameterTypeRef=\"peaksat-dt/{variable_name}_t\" name=\"{acronym}{variable_name}\"")
    else:
        parameter_lines.append(f"            <Parameter parameterTypeRef=\"base-dt/{variable_type}\" name=\"{acronym}{variable_name}\"")

    enum_lines = []
    if variable_type == "enum":
        enum_lines.append(f'')
        enum_lines.append(f'            <EnumeratedParameterType name="{variable_name}_t">')
        enum_lines.append('                <IntegerDataEncoding sizeInBits="8" />')
        enum_lines.append('                <EnumerationList>')
        for enum_item in enum_items.split(","):
            enum_item = enum_item.strip()
            if not enum_item:
                continue
            enum_item_parts = enum_item.split(" = ")
            if len(enum_item_parts) == 2:
                label, value = enum_item_parts
                enum_lines.append(f'                    <Enumeration label="{label.strip()}" value="{value.strip()}" />')
        enum_lines.append('                </EnumerationList>')
        enum_lines.append('            </EnumeratedParameterType>')
        enum_lines.append(f'')

    return parameter_lines, enum_lines

def finalize_output_files(dt_lines, xtce_lines, parameter_id_lines):
    dt_lines.append('        </ParameterTypeSet>')
    dt_lines.append('    </xtce:TelemetryMetaData>')
    dt_lines.append('</SpaceSystem>')

    xtce_lines.append('        </ParameterSet>\n')
    xtce_lines.append('    </xtce:TelemetryMetaData>\n')
    xtce_lines.append('</SpaceSystem>')

    id_lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<SpaceSystem name="dt" xmlns:xtce="http://www.omg.org/spec/XTCE/20180204"',
        '    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"',
        '    xsi:schemaLocation="http://www.omg.org/spec/XTCE/20180204 https://www.omg.org/spec/XTCE/20180204/SpaceSystem.xsd"',
        '    shortDescription="This is a bogus satellite telemetry and telecommand database."',
        '    operationalStatus="unittest">',
        '<!--Contains all ParameterTypes for Telemetry and ArgumentTypes for Telecommanding.-->',
        '    <TelemetryMetaData>',
        '        <ParameterTypeSet>',
        '            <EnumeratedParameterType name="parameterId_t">',
        '                <IntegerDataEncoding sizeInBits="16" />',
        '                <EnumerationList>',
        *parameter_id_lines,
        '                </EnumerationList>',
        '            </EnumeratedParameterType>',
        ' ',
        '            <ArrayParameterType name="parameterIdArray_t" arrayTypeRef="parameterId_t">',
        '               <DimensionList>',
        '                   <Dimension>',
        '                       <StartingIndex>',
        '                           <FixedValue>0</FixedValue>',
        '                      </StartingIndex>',
        '                      <EndingIndex>',
        '                           <DynamicValue>',
        '                               <ArgumentInstanceRef argumentRef="total_parameters" />',
        '                               <LinearAdjustment intercept="-1" />',
        '                           </DynamicValue>',
        '                       </EndingIndex>',
        '                   </Dimension>',
        '               </DimensionList>',
        '           </ArrayParameterType>',
        '        </ParameterTypeSet>',
        '    </TelemetryMetaData>',
        ' ',
        '    <CommandMetaData>',
        '        <ArgumentTypeSet>',
        '            <EnumeratedArgumentType name="parameterId_t">',
        '                <IntegerDataEncoding sizeInBits="16" />',
        '                <EnumerationList>',
        *parameter_id_lines,
        '                </EnumerationList>',
        '            </EnumeratedArgumentType>',
        '        </ArgumentTypeSet>',
        ' ',
        '       <ArgumentTypeSet>',
        '           <ArrayArgumentType name="parameterIdArray_t" arrayTypeRef="parameterId_t">',
        '              <DimensionList>',
        '                  <Dimension>',
        '                      <StartingIndex>',
        '                          <FixedValue>0</FixedValue>',
        '                     </StartingIndex>',
        '                     <EndingIndex>',
        '                          <DynamicValue>',
        '                              <ArgumentInstanceRef argumentRef="total_parameters" />',
        '                              <LinearAdjustment intercept="-1" />',
        '                          </DynamicValue>',
        '                      </EndingIndex>',
        '                  </Dimension>',
        '              </DimensionList>',
        '          </ArrayArgumentType>',
        '      </ArgumentTypeSet>',
        '   </CommandMetaData>',
        '</SpaceSystem>'
    ]

    output_dir = "src/main/yamcs/mdb/peaksat/"
    os.makedirs(output_dir, exist_ok=True)

    output_xtce_file = f"{output_dir}peaksat-xtce.xml"
    output_dt_file = f"{output_dir}peaksat-dt.xml"
    output_parameter_id_file = f"{output_dir}parameter_id.xml"

    with open(output_xtce_file, "w") as xtce_file:
        xtce_file.write("\n".join(xtce_lines))

    with open(output_dt_file, "w") as dt_file:
        dt_file.write("\n".join(line for line in dt_lines if line.strip()))

    with open(output_parameter_id_file, "w") as parameter_id_file:
        parameter_id_file.write("\n".join(id_lines))

    print(f"Processing complete.")
    print(f"Generated XTCE file: {output_xtce_file}")
    print(f"Generated DT file: {output_dt_file}")
    print(f"Generated Parameter ID file: {output_parameter_id_file}")

# Main execution
spreadsheet_id = "12m4Sq4CMnLB9nn6INxKUdSNSIlTJes_uTbg19RE9-X0"
xlsx_export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
output_file = "mission_database.xlsx"
download_excel_file(xlsx_export_url, output_file)

type_encoding = {
    "uint8_t": 0,
    "bool_t" : 0,
    "int8_t": 1,
    "uint16_t": 2,
    "int16_t": 3,
    "uint32_t": 4,
    "int32_t": 5,
    "uint64_t": 6,
    "int64_t": 7,
    "float_t": 8,
    "double_t": 9,
}

subsystem_config = {
    "OBDH": 5000,
    "COMMS": 10000,
    "PAY": 15000,
    "ADCS": 20000,
    "EPS": 25000,
}

cpp_type_map = {
    "bool": "bool_t",
    "float": "float_t",
    "double": "double_t",
    "int": "int_t",
    "uint8_t": "uint8_t",
    "uint16_t": "uint16_t",
    "uint32_t": "uint32_t",
    "uint64_t": "uint64_t",
    "int8_t": "int8_t",
    "int16_t": "int16_t",
    "int32_t": "int32_t",
    "int64_t": "int64_t",
    "char": "char_t",
    "enum": "enum",
}

process_excel_file(output_file, subsystem_config, cpp_type_map, type_encoding)